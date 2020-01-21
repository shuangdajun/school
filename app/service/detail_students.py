from datetime import datetime

import xlrd, xlwt
import re

import openpyxl
from xlrd import xldate_as_tuple

from app.model.Base import db
from app.model.Prices import Prices
from app.model.Students import Students, to_stu_sub
from app.model.Subjects import Subjects
from app.model.Teachers import Teachers


from app.form import judgeStuentForm, judgeTeacherForm, judgeSubjectForm

field_student=["student_name","student_sex","student_age","student_phone","student_landline"]
field_teacher=["teacher_name","teacher_address","teacher_phone"]
field_subject=["subject_name","sub_stu","sub_tea"]
field_prices=["student_name","teacher_name","subject_name","startTime","stopTime","ClassHours","prices"]

def if_time(time,wb):
    if time==None:
        return  ""
    return  datetime(*xldate_as_tuple(time, wb.datemode)).strftime("%Y-%m-%d")
def if_str(Object,str):
    if Object==None:
        return ""
    return getattr(Object,str)
#bytes转换为字符串数组
def bytes_str_set(data):
    data_new=[]
    for value in data:
        data_new.append(value.decode("utf-8"))
    return data_new
def judge_add_student_form(form_dict):
    judge_formclass = judgeStuentForm(form_dict)
    judge = judge_formclass.validate()

    return [judge, judge_formclass]


def judge_add_teacher_form(form_dict):
    judge_formclass = judgeTeacherForm(form_dict)
    judge = judge_formclass.validate()
    return [judge, judge_formclass]


def judge_add_subject_form(form_dict):
    judge_formclass = judgeSubjectForm(form_dict)
    judge = judge_formclass.validate()
    return [judge, judge_formclass]


def xlsx_upload(file, path):
    if file is None:
        return False
    with open(path, "wb") as f:
        data = file.read()
        f.write(data)
        f.close()
    return True

def xlsx_upload_student(path):

    wb=xlrd.open_workbook(path)
    sheet=wb.sheet_by_name("Sheet1")

    for row in range(1,sheet.nrows):
        try:
            student = Students()
            student.student_name=sheet.cell_value(row,0)
            student.student_sex=sheet.cell_value(row,1)
            student.student_age=sheet.cell_value(row,2)
            student.student_phone = sheet.cell_value(row, 3)

        except Exception as e:
            print(e)
        db.session.add(student)
    try:
        db.session.commit()
    except Exception as e:
        print(e)
def xlsx_upload_teacher(path):

    wb=xlrd.open_workbook(path)
    sheet=wb.sheet_by_name("Sheet1")
    for row in range(1,sheet.nrows):
        teacher = Teachers()

        teacher.teacher_name=sheet.cell_value(row,0)
        teacher.position=sheet.cell_value(row,1)
        teacher.teacher_phone=sheet.cell_value(row,2)

        teacher.teacher_address=sheet.cell_value(row,3)
        db.session.add(teacher)
    try:
        db.session.commit()
    except Exception as e:
        print(e)

def xlsx_upload_subject(path):

    wb=xlrd.open_workbook(path)
    sheet=wb.sheet_by_name("Sheet1")

    for row in range(1,sheet.nrows):
        subject = Subjects()
        subject.subject_name=sheet.cell_value(row,0)
        subject.class_name = sheet.cell_value(row, 1)
        stuNameList=sheet.cell_value(row,2).split(",")
        studentList=[]
        for student_name in stuNameList:
            student=Students.query.filter_by(student_name=student_name).first()
            if student:
                studentList.append(student)
        subject.sub_stu=studentList

        teaNameList = sheet.cell_value(row,3).split(",")
        teacherList = []
        for teacher_name in teaNameList:
            teacher = Teachers.query.filter_by(teacher_name=teacher_name).first()
            if teacher:
                teacherList.append(teacher)
        subject.sub_tea = teacherList


        db.session.add(subject)
    try:
        db.session.commit()
    except Exception as e:
        print(e)

def xlsx_upload_price(path):
    wb=xlrd.open_workbook(path)
    sheet=wb.sheet_by_name("Sheet1")

    for row in range(1, sheet.nrows):
        try:
            price = Prices()
            price.pri_stu=Students.query.filter(Students.student_name==sheet.cell_value(row,0)).first()
            price.pri_tea = Teachers.query.filter(Teachers.teacher_name==sheet.cell_value(row, 1)).first()
            price.pri_sub = Subjects.query.filter(Subjects.subject_name==sheet.cell_value(row, 2)).first()

            price.startTime = if_time(sheet.cell_value(row,3),wb)
            price.stopTime=  if_time(sheet.cell_value(row,4),wb)
            price.ClassHours=sheet.cell_value(row,5)
            price.prices=sheet.cell_value(row,6)
            db.session.add(price)
        except Exception as e:
            print(e)
    try:
        db.session.commit()
    except Exception as e:
        print(e)



def export_file(filename):
    if re.search("Students_info", filename):
        result = export_db_stu_tea(Students, filename,field_student)
    elif re.search("Teachers_info", filename):
        result = export_db_stu_tea(Teachers, filename,field_teacher )
    elif re.search("Subjects_info", filename):
        result = export_db_subject( filename,field_subject)
    elif re.search("Price_info",filename):
        result = export_db_price(filename,field_prices)
    else:
        return False
    return result


def export_db_stu_tea(Object, path,field):
    workbook = xlwt.Workbook(encoding="utf-8")
    sheet = workbook.add_sheet("Sheet1")
    students = Object.query.all()
    rows = len(students)

    try:
        for col in range(0, len(field)):
            sheet.write(0, col, field[col])

        for row in range(1, rows + 1):
            for col in range(0, len(field)):
                        sheet.write(row, col, getattr(students[row - 1], field[col]))


        workbook.save(path + ".xlsx")
    except Exception as e:
        return False
    return True


def export_db_subject(path, field):
    workbook = xlwt.Workbook(encoding="utf-8")
    sheet = workbook.add_sheet("Sheet1")
    subjects = Subjects.query.all()
    rows = len(subjects)


    try:
        for col in range(0, len(field)):
            if field[col] == "sub_stu":
                sheet.write(0, col, "student_name")
            elif field[col] == "sub_tea":
                sheet.write(0, col, "teacher_name")
            else:
                sheet.write(0, col, field[col])

        for row in range(1, rows + 1):
            for col in range(0, len(field)):
                data = ""
                if field[col] == "sub_stu":
                    for student in getattr(subjects[row - 1], field[col]):
                        data=(data+","+student.student_name).strip(",")

                    sheet.write(row, col, data)
                elif field[col] == "sub_tea":
                    for teacher in getattr(subjects[row - 1], field[col]):
                        data = (data + "," + teacher.teacher_name).strip(",")
                    sheet.write(row, col, data)
                else:
                    sheet.write(row, col, getattr(subjects[row - 1], field[col]))
        workbook.save(path + ".xlsx")
    except Exception as e:
        return False
    return True
def export_db_price(path,field):
    workbook = xlwt.Workbook(encoding="utf-8")
    sheet = workbook.add_sheet("Sheet1")
    prices = Prices.query.all()
    rows = len(prices)


    try:
        for col in range(0, len(field)):
                sheet.write(0, col, field[col])

        for row in range(1, rows + 1):
            for col in range(0, len(field)):

                if field[col] == "student_name":
                    sheet.write(row, col, if_str(prices[row-1].pri_stu,"student_name"))
                elif field[col] == "teacher_name":
                    sheet.write(row, col, if_str(prices[row - 1].pri_tea, "teacher_name"))

                elif field[col] == "subject_name":
                    sheet.write(row, col, if_str(prices[row - 1].pri_sub, "subject_name"))
                else:
                    sheet.write(row, col, getattr(prices[row - 1], field[col]))
        workbook.save(path + ".xlsx")
    except Exception as e:
        return False
    return True

def getObjList(ObjList, attr):
    ObjAttrList = []
    for Obj in ObjList:
        ObjAttrList.append(getattr(Obj, attr))
    return ObjAttrList


def SubStuCountDict(subjects,stuCount):
    stuCount=len(subjects.sub_stu)
    return {"subject_name":subjects.subject_name,"student_count":stuCount,"percentage":stuCount/stuCount}



def dict_if(param_dict):
    result={}
    keyslist=list(param_dict.keys())

    for i in range((len(param_dict)//2)):
        if param_dict[keyslist[i*2]] and param_dict[keyslist[i*2+1]]:
            result[param_dict[keyslist[i*2]]]=param_dict[keyslist[i*2+1]]

    return result