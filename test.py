# -*- coding:utf-8
from sqlalchemy import Column, String, Integer, create_engine, ForeignKey, MetaData, Table
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, relationship
from openpyxl import load_workbook
wb=load_workbook("C:\\Users\\x\\Documents\\WeChat Files\\wxid_vbfe2gu2k0rc22\FileStorage\\File\\2019-11\\学生表(2)(1).xlsx")
sheet_names=wb.sheetnames
for sheet_name in sheet_names:
    ws=wb[sheet_name]
    max_row=ws.max_row
    max_col=ws.max_column
    list=""
    for row in range(2,max_row+1):
        if ws.cell(row=row,column=2).value==None:
            continue
        list=list+","+ws.cell(row=row,column=2).value
    print(sheet_name+" "+list)